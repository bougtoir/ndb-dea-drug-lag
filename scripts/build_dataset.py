#!/usr/bin/env python3
"""
Build the HCV drug-lag time series from NDB Open Data (editions 1-10, FY2014-FY2023).

Reads the raw NDB "処方薬 性年齢別薬効分類別数量" workbooks (内服/外用/注射,
外来院内・外来院外・入院) downloaded under data/ndb_raw/dai{N}/f00..f04.xlsx and
extracts the national total dispensed quantity (総計 / 処方数量) for the
hepatitis-C direct-acting antivirals (DAAs) and the interferon-based backbone
(peginterferon, conventional interferon, ribavirin).

Output:
  data/target_drugs_long.csv     one row per (edition, product, formulation, setting)
  data/hcv_timeseries.csv        group x fiscal-year national totals
  data/hcv_product_timeseries.csv product x fiscal-year national totals

Metric note: NDB Open Data reports 処方数量 (dispensed quantity: tablets/capsules
for oral drugs, pre-filled syringes/vials for injections), NOT patient counts.
Quantities are therefore comparable within a product/formulation over time but are
not additive across products with different dosage units. Estimated treatment
courses (an approximate patient-count proxy) are derived separately in analyze.py
using documented regimen durations, and are clearly labelled as estimates.
"""
import glob
import os

import openpyxl
import pandas as pd

ROOT = os.path.join(os.path.dirname(__file__), "..", "data", "ndb_raw")
OUT = os.path.join(os.path.dirname(__file__), "..", "data")

EDITION_FY = {1: 2014, 2: 2015, 3: 2016, 4: 2017, 5: 2018,
              6: 2019, 7: 2020, 8: 2021, 9: 2022, 10: 2023}

# Interferon-free DAA products (brand-name substrings as they appear in NDB) -> INN label
DAA = {
    "ソバルディ": "sofosbuvir",
    "ハーボニー": "ledipasvir/sofosbuvir",
    "ダクルインザ": "daclatasvir",
    "スンベプラ": "asunaprevir",
    "マヴィレット": "glecaprevir/pibrentasvir",
    "エプクルーサ": "sofosbuvir/velpatasvir",
    "エレルサ": "elbasvir",
    "グラジナ": "grazoprevir",
    "ヴィキラックス": "ombitasvir/paritaprevir/ritonavir",
}
# First-generation NS3/4A protease inhibitors used *with* peginterferon+ribavirin
# (interferon-BASED triple therapy, not interferon-free). Kept separate so the
# DAA group cleanly represents interferon-free regimens.
PI_IFN = {
    "テラビック": "telaprevir",
    "ソブリアード": "simeprevir",
    "バニヘップ": "vaniprevir",
}
IFN_PEG = ["ペガシス", "ペグイントロン"]
IFN_CONV = ["スミフェロン", "フエロン", "イントロンＡ", "オーアイエフ"]
RIBAVIRIN = ["レベトール", "コペガス"]


def classify(name: str):
    for k, v in DAA.items():
        if k in name:
            return "DAA", v
    for k, v in PI_IFN.items():
        if k in name:
            return "PI_ifn", v
    if any(k in name for k in IFN_PEG):
        return "IFN_peg", "peginterferon"
    if any(k in name for k in IFN_CONV):
        return "IFN_conv", "interferon_conventional"
    if any(k in name for k in RIBAVIRIN):
        return "ribavirin", "ribavirin"
    return None, None


def header_cols(ws):
    """Locate header row and column indices; NDB layout varies across editions."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True)):
        vals = [str(c).replace("\n", "") if c is not None else "" for c in row]
        if "薬効分類名称" in vals and any(v.startswith("総計") for v in vals):
            idx = {}
            for j, v in enumerate(vals):
                if v == "薬効分類名称":
                    idx["cls"] = j
                elif v == "医薬品名":
                    idx["name"] = j
                elif v == "単位":
                    idx["unit"] = j
                elif v.startswith("総計"):
                    idx["tot"] = j
            return i + 1, idx
    return None, None


def main():
    rows = []
    for ed in range(1, 11):
        # Read every workbook for the edition (not just f00-f04): additional
        # medical/dental supplement files are harmless (they contain none of the
        # target drugs) but scanning them all avoids silently dropping any file
        # if the edition's file layout changes.
        for fn in sorted(glob.glob(os.path.join(ROOT, f"dai{ed}", "f*.xlsx"))):
            wb = openpyxl.load_workbook(fn, read_only=True)
            for sh in wb.sheetnames:
                ws = wb[sh]
                hrow, idx = header_cols(ws)
                if not idx:
                    raise RuntimeError(f"header not found: dai{ed} {sh}")
                cur_cls = None
                for r in ws.iter_rows(min_row=hrow + 1, values_only=True):
                    cls = r[idx["cls"]] if idx["cls"] < len(r) else None
                    if cls:
                        cur_cls = cls
                    nm = r[idx["name"]] if idx["name"] < len(r) else None
                    if not nm:
                        continue
                    grp, prod = classify(str(nm))
                    if grp is None:
                        continue
                    unit = r[idx["unit"]] if "unit" in idx and idx["unit"] < len(r) else ""
                    tot = r[idx["tot"]] if idx["tot"] < len(r) else None
                    try:
                        tot = float(tot)
                    except (TypeError, ValueError):
                        tot = 0.0
                    rows.append([ed, EDITION_FY[ed], grp, prod, str(nm),
                                 str(cur_cls), unit, sh, tot])
            wb.close()

    df = pd.DataFrame(rows, columns=["edition", "fy", "group", "product", "drug_name",
                                     "yakko_class", "unit", "sheet", "total_qty"])
    df.to_csv(os.path.join(OUT, "target_drugs_long.csv"), index=False, encoding="utf-8-sig")

    grp_ts = (df.groupby(["group", "fy"])["total_qty"].sum()
              .reset_index()
              .pivot(index="fy", columns="group", values="total_qty")
              .fillna(0.0))
    grp_ts.to_csv(os.path.join(OUT, "hcv_timeseries.csv"), encoding="utf-8-sig")

    prod_ts = (df.groupby(["group", "product", "fy"])["total_qty"].sum()
               .reset_index()
               .pivot_table(index=["group", "product"], columns="fy",
                            values="total_qty", fill_value=0.0))
    prod_ts.to_csv(os.path.join(OUT, "hcv_product_timeseries.csv"), encoding="utf-8-sig")

    print("rows:", len(df))
    print(grp_ts.round(0))
    return df


if __name__ == "__main__":
    main()
